"""
Excel Parser Module

Handles parsing of Excel files (.xlsx, .xlsm) to extract:
- Worksheets and their structure
- Cell values and formulas
- Named ranges
- Data tables and pivot tables
- Metadata and styles
"""

from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from loguru import logger


@dataclass
class CellInfo:
    """Information about a single cell."""
    sheet_name: str
    address: str  # e.g., "A1"
    row: int
    col: int
    value: Any
    formula: Optional[str] = None
    data_type: Optional[str] = None
    number_format: Optional[str] = None
    is_merged: bool = False
    comment: Optional[str] = None


@dataclass
class NamedRange:
    """Information about a named range."""
    name: str
    sheet_name: Optional[str]
    address: str  # e.g., "Sheet1!$A$1:$A$10"
    scope: str  # "workbook" or sheet name


@dataclass
class SheetInfo:
    """Information about a worksheet."""
    name: str
    index: int
    visible: bool
    max_row: int
    max_col: int
    cells: Dict[str, CellInfo] = field(default_factory=dict)
    tables: List[str] = field(default_factory=list)
    pivot_tables: List[str] = field(default_factory=list)


@dataclass
class ExcelStructure:
    """Complete structure of an Excel workbook."""
    filename: str
    sheets: Dict[str, SheetInfo]
    named_ranges: Dict[str, NamedRange]
    has_vba: bool
    properties: Dict[str, Any] = field(default_factory=dict)


class ExcelParser:
    """
    Parse Excel files to extract structure, data, and formulas.
    
    Uses openpyxl for deterministic parsing of Excel file structure.
    """
    
    def __init__(self, data_only: bool = False):
        """
        Initialize the Excel parser.
        
        Args:
            data_only: If True, read cell values instead of formulas
        """
        self.data_only = data_only
        self.workbook: Optional[Workbook] = None
        self.structure: Optional[ExcelStructure] = None
        
    def parse(self, filepath: str, password: Optional[str] = None) -> ExcelStructure:
        """
        Parse an Excel file and extract its complete structure.
        
        Args:
            filepath: Path to the Excel file
            password: Password for protected files
            
        Returns:
            ExcelStructure containing all parsed information
        """
        logger.info(f"Parsing Excel file: {filepath}")
        
        file_path = Path(filepath)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {filepath}")
        
        # Load workbook
        try:
            self.workbook = openpyxl.load_workbook(
                filepath,
                data_only=self.data_only,
                keep_vba=True  # Preserve VBA for .xlsm files
            )
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise
        
        # Check for VBA
        has_vba = file_path.suffix.lower() == '.xlsm' and hasattr(self.workbook, 'vba_archive')
        
        # Initialize structure
        self.structure = ExcelStructure(
            filename=file_path.name,
            sheets={},
            named_ranges={},
            has_vba=has_vba
        )
        
        # Parse workbook properties
        self._parse_properties()
        
        # Parse all worksheets
        self._parse_worksheets()
        
        # Parse named ranges
        self._parse_named_ranges()
        
        logger.info(f"Successfully parsed {len(self.structure.sheets)} sheets")
        return self.structure
    
    def _parse_properties(self):
        """Extract workbook properties and metadata."""
        if not self.workbook:
            return
            
        props = self.workbook.properties
        self.structure.properties = {
            'creator': props.creator,
            'title': props.title,
            'subject': props.subject,
            'description': props.description,
            'created': props.created,
            'modified': props.modified,
        }
        
    def _parse_worksheets(self):
        """Parse all worksheets in the workbook."""
        if not self.workbook:
            return
            
        for idx, sheet in enumerate(self.workbook.worksheets):
            logger.debug(f"Parsing sheet: {sheet.title}")
            
            sheet_info = SheetInfo(
                name=sheet.title,
                index=idx,
                visible=sheet.sheet_state == 'visible',
                max_row=sheet.max_row,
                max_col=sheet.max_column
            )
            
            # Parse cells
            self._parse_cells(sheet, sheet_info)
            
            # Parse tables
            if hasattr(sheet, 'tables'):
                sheet_info.tables = list(sheet.tables.keys())
            
            # Parse pivot tables
            if hasattr(sheet, '_pivots'):
                sheet_info.pivot_tables = [p.name for p in sheet._pivots]
            
            self.structure.sheets[sheet.title] = sheet_info
    
    def _parse_cells(self, sheet: Worksheet, sheet_info: SheetInfo):
        """
        Parse all cells in a worksheet.
        
        Args:
            sheet: The worksheet to parse
            sheet_info: SheetInfo object to populate
        """
        # Iterate through all cells with content
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                   min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is None and not cell.data_type:
                    continue  # Skip empty cells
                
                # Get cell address
                address = cell.coordinate
                
                # Extract formula if present
                formula = None
                if cell.data_type == 'f':  # Formula
                    formula = cell.value if isinstance(cell.value, str) else None
                
                # Check if cell is part of merged range
                is_merged = any(address in merged_range for merged_range in sheet.merged_cells.ranges)
                
                # Extract comment
                comment = cell.comment.text if cell.comment else None
                
                # Create CellInfo object
                cell_info = CellInfo(
                    sheet_name=sheet.title,
                    address=address,
                    row=cell.row,
                    col=cell.column,
                    value=cell.value,
                    formula=formula,
                    data_type=cell.data_type,
                    number_format=cell.number_format,
                    is_merged=is_merged,
                    comment=comment
                )
                
                sheet_info.cells[address] = cell_info
    
    def _parse_named_ranges(self):
        """Parse all named ranges in the workbook."""
        if not self.workbook:
            return
        
        for name, defn in self.workbook.defined_names.items():
            # Parse the definition to extract sheet and address
            destinations = list(defn.destinations)
            
            if destinations:
                sheet_name, address = destinations[0]
                
                named_range = NamedRange(
                    name=name,
                    sheet_name=sheet_name,
                    address=f"{sheet_name}!{address}" if sheet_name else address,
                    scope=defn.localSheetId if hasattr(defn, 'localSheetId') else 'workbook'
                )
                
                self.structure.named_ranges[name] = named_range
    
    def get_sheet(self, sheet_name: str) -> Optional[SheetInfo]:
        """Get information about a specific sheet."""
        if not self.structure:
            return None
        return self.structure.sheets.get(sheet_name)
    
    def get_cell(self, sheet_name: str, address: str) -> Optional[CellInfo]:
        """Get information about a specific cell."""
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None
        return sheet.cells.get(address)
    
    def get_formulas(self, sheet_name: Optional[str] = None) -> List[CellInfo]:
        """
        Get all cells containing formulas.
        
        Args:
            sheet_name: If specified, only return formulas from this sheet
            
        Returns:
            List of CellInfo objects for cells with formulas
        """
        if not self.structure:
            return []
        
        formulas = []
        sheets_to_check = [self.structure.sheets[sheet_name]] if sheet_name else self.structure.sheets.values()
        
        for sheet in sheets_to_check:
            for cell in sheet.cells.values():
                if cell.formula:
                    formulas.append(cell)
        
        return formulas
    
    def get_named_range_cells(self, range_name: str) -> List[CellInfo]:
        """
        Get all cells in a named range.
        
        Args:
            range_name: Name of the range
            
        Returns:
            List of CellInfo objects in the range
        """
        if not self.structure or range_name not in self.structure.named_ranges:
            return []
        
        named_range = self.structure.named_ranges[range_name]
        # TODO: Parse address range and return cells
        # This requires parsing addresses like "Sheet1!$A$1:$A$10"
        return []
    
    def export_structure(self) -> Dict[str, Any]:
        """
        Export the parsed structure as a dictionary.
        
        Returns:
            Dictionary representation of the Excel structure
        """
        if not self.structure:
            return {}
        
        return {
            'filename': self.structure.filename,
            'has_vba': self.structure.has_vba,
            'properties': self.structure.properties,
            'sheets': {
                name: {
                    'index': sheet.index,
                    'visible': sheet.visible,
                    'max_row': sheet.max_row,
                    'max_col': sheet.max_col,
                    'cell_count': len(sheet.cells),
                    'formula_count': sum(1 for c in sheet.cells.values() if c.formula),
                    'tables': sheet.tables,
                    'pivot_tables': sheet.pivot_tables
                }
                for name, sheet in self.structure.sheets.items()
            },
            'named_ranges': {
                name: {
                    'address': nr.address,
                    'scope': nr.scope
                }
                for name, nr in self.structure.named_ranges.items()
            }
        }
    
    def close(self):
        """Close the workbook and free resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
